"""
User Activity Logging Module
Maintains a local Excel file (user_activity_logs.xlsx) with user access and activity data.
Thread-safe with batched writes for minimal I/O overhead.
"""

import os
import uuid
import shutil
import threading
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# === Configuration ===
BASE_DIR = Path(__file__).resolve().parent.parent  # Backend/
LOG_FILE = BASE_DIR / "user_activity_logs.xlsx"
BACKUP_DIR = BASE_DIR / "backups"

# Excel columns (order matters)
COLUMNS = [
    "user_id",
    "username",
    "email",
    "last_login",
    "session_start",
    "session_end",
    "session_id",
    "total_queries",
    "features_used",
    "last_activity",
]


class ActivityLogger:
    """
    Thread-safe Excel-based activity logger.
    Singleton pattern — use ActivityLogger.get_instance().
    """

    _instance = None
    _lock = threading.Lock()

    def __init__(self):
        self._file_lock = threading.Lock()
        self._write_buffer = []
        self._buffer_limit = 10
        self._ensure_file_exists()
        print(f"[ActivityLogger] Initialized. Log file: {LOG_FILE}")

    @classmethod
    def get_instance(cls) -> "ActivityLogger":
        """Thread-safe singleton accessor."""
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = cls()
        return cls._instance

    # ─── File Management ────────────────────────────────────────

    def _ensure_file_exists(self):
        """Create the Excel file with headers if it doesn't exist."""
        if LOG_FILE.exists():
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Activity Logs"

        # Style the header row
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid")
        thin_border = Border(
            bottom=Side(style="thin", color="CCCCCC")
        )

        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            # Auto-width hint
            ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 18)

        wb.save(str(LOG_FILE))
        wb.close()
        print(f"[ActivityLogger] Created new log file: {LOG_FILE}")

    def _find_user_row(self, ws, user_id: str) -> Optional[int]:
        """Find the row number for a given user_id. Returns None if not found."""
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == user_id:
                return row
        return None

    def _col_index(self, col_name: str) -> int:
        """Get 1-based column index for a column name."""
        return COLUMNS.index(col_name) + 1

    # ─── Logging Methods ────────────────────────────────────────

    def log_login(self, user_id: str, username: str, email: str) -> str:
        """
        Log a user login event. Creates or updates the user's row.
        Returns the new session_id.
        """
        session_id = str(uuid.uuid4())[:8]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with self._file_lock:
            try:
                wb = openpyxl.load_workbook(str(LOG_FILE))
                ws = wb.active

                row = self._find_user_row(ws, user_id)

                if row:
                    # Update existing user
                    ws.cell(row=row, column=self._col_index("username"), value=username)
                    ws.cell(row=row, column=self._col_index("email"), value=email)
                    ws.cell(row=row, column=self._col_index("last_login"), value=now)
                    ws.cell(row=row, column=self._col_index("session_start"), value=now)
                    ws.cell(row=row, column=self._col_index("session_end"), value="")
                    ws.cell(row=row, column=self._col_index("session_id"), value=session_id)
                    ws.cell(row=row, column=self._col_index("last_activity"), value=now)
                else:
                    # New user row
                    new_row = [
                        user_id,
                        username,
                        email,
                        now,           # last_login
                        now,           # session_start
                        "",            # session_end
                        session_id,
                        0,             # total_queries
                        "",            # features_used
                        now,           # last_activity
                    ]
                    ws.append(new_row)

                wb.save(str(LOG_FILE))
                wb.close()
                print(f"[ActivityLogger] Login logged for {username} (session: {session_id})")
            except Exception as e:
                print(f"[ActivityLogger] ERROR logging login: {e}")

        return session_id

    def log_query(self, user_id: str, feature: str = "QUERY"):
        """
        Log a query event. Increments total_queries and updates features_used.
        """
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with self._file_lock:
            try:
                wb = openpyxl.load_workbook(str(LOG_FILE))
                ws = wb.active

                row = self._find_user_row(ws, user_id)
                if not row:
                    # User hasn't logged in yet — skip silently
                    wb.close()
                    return

                # Increment query count
                current_count = ws.cell(row=row, column=self._col_index("total_queries")).value or 0
                ws.cell(row=row, column=self._col_index("total_queries"), value=current_count + 1)

                # Update features used
                current_features = ws.cell(row=row, column=self._col_index("features_used")).value or ""
                feature_set = set(f.strip() for f in current_features.split(",") if f.strip())
                feature_set.add(feature.upper())
                ws.cell(row=row, column=self._col_index("features_used"), value=", ".join(sorted(feature_set)))

                # Update last activity
                ws.cell(row=row, column=self._col_index("last_activity"), value=now)

                wb.save(str(LOG_FILE))
                wb.close()
                print(f"[ActivityLogger] Query logged for user {user_id} (feature: {feature})")
            except Exception as e:
                print(f"[ActivityLogger] ERROR logging query: {e}")

    def log_logout(self, user_id: str):
        """
        Log a session end for the user.
        """
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with self._file_lock:
            try:
                wb = openpyxl.load_workbook(str(LOG_FILE))
                ws = wb.active

                row = self._find_user_row(ws, user_id)
                if row:
                    ws.cell(row=row, column=self._col_index("session_end"), value=now)
                    ws.cell(row=row, column=self._col_index("last_activity"), value=now)
                    wb.save(str(LOG_FILE))

                wb.close()
                print(f"[ActivityLogger] Logout logged for user {user_id}")
            except Exception as e:
                print(f"[ActivityLogger] ERROR logging logout: {e}")

    # ─── Export & Backup ────────────────────────────────────────

    def get_all_logs(self) -> list:
        """Read all logs and return as list of dicts."""
        with self._file_lock:
            try:
                wb = openpyxl.load_workbook(str(LOG_FILE))
                ws = wb.active

                logs = []
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row=row, column=1).value is None:
                        continue
                    entry = {}
                    for col_idx, col_name in enumerate(COLUMNS, 1):
                        val = ws.cell(row=row, column=col_idx).value
                        entry[col_name] = str(val) if val is not None else ""
                    logs.append(entry)

                wb.close()
                return logs
            except Exception as e:
                print(f"[ActivityLogger] ERROR reading logs: {e}")
                return []

    def backup(self) -> str:
        """
        Create a timestamped backup of the log file.
        Returns the backup file path.
        """
        BACKUP_DIR.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = BACKUP_DIR / f"user_activity_logs_backup_{timestamp}.xlsx"

        with self._file_lock:
            try:
                shutil.copy2(str(LOG_FILE), str(backup_path))
                print(f"[ActivityLogger] Backup created: {backup_path}")
                return str(backup_path)
            except Exception as e:
                print(f"[ActivityLogger] ERROR creating backup: {e}")
                return ""

    def get_file_path(self) -> str:
        """Return the absolute path to the log file."""
        return str(LOG_FILE)


# Convenience: module-level singleton access
def get_logger() -> ActivityLogger:
    return ActivityLogger.get_instance()
